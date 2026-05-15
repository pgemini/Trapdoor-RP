"""Excel spreadsheet extractor (.xlsx / .xlsm).

Pulls every non-empty cell, every cell comment, defined names, and the
workbook's core properties. Formulas are preserved verbatim because they
are themselves a known injection / data-exfil vector (e.g. HYPERLINK,
WEBSERVICE) and we want detectors to see the raw expression.
"""
from __future__ import annotations

import io

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..schemas import ExtractedContent, ExtractedFragment
from .base import Extractor


# Cell limit per sheet — protects against multi-million-row workbooks.
_MAX_CELLS_PER_SHEET = 5_000


class ExcelExtractor(Extractor):
    modality = "spreadsheet/xlsx"

    def extract(self, data: bytes, filename: str) -> ExtractedContent:
        wb = load_workbook(io.BytesIO(data), data_only=False, read_only=False)
        fragments: list[ExtractedFragment] = []
        notes: list[str] = []
        sheet_count = len(wb.sheetnames)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            seen = 0
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if seen >= _MAX_CELLS_PER_SHEET:
                        break
                    value = cell.value
                    if value is None or value == "":
                        continue
                    cell_ref = f"{sheet_name}!{get_column_letter(cell.column)}{cell.row}"
                    text = str(value)
                    fragments.append(ExtractedFragment(
                        source=f"cell.{cell_ref}",
                        kind="text",
                        text=text,
                        attrs={"is_formula": text.startswith("=")},
                    ))
                    if cell.comment:
                        c_text = (cell.comment.text or "").strip()
                        if c_text:
                            fragments.append(ExtractedFragment(
                                source=f"comment.{cell_ref}",
                                kind="comment",
                                text=c_text,
                                attrs={"author": cell.comment.author or ""},
                            ))
                    seen += 1
                if seen >= _MAX_CELLS_PER_SHEET:
                    notes.append(f"Sheet '{sheet_name}' truncated at {_MAX_CELLS_PER_SHEET} cells.")
                    break

        # Defined names (named ranges + their formulas) — another sneaky path.
        try:
            for name in wb.defined_names:
                dn = wb.defined_names[name]
                if dn and dn.value:
                    fragments.append(ExtractedFragment(
                        source=f"named.{name}",
                        kind="metadata",
                        text=str(dn.value),
                    ))
        except Exception:
            pass

        # Workbook properties — same playbook as DOCX metadata injection.
        props = wb.properties
        meta_keys = ("title", "subject", "creator", "description",
                     "keywords", "category", "lastModifiedBy", "contentStatus",
                     "identifier", "language")
        meta_dump: dict[str, str] = {}
        for key in meta_keys:
            value = getattr(props, key, None)
            if value:
                text = str(value)
                meta_dump[key] = text
                fragments.append(ExtractedFragment(
                    source=f"metadata.{key}",
                    kind="metadata",
                    text=text,
                ))

        return ExtractedContent(
            modality=self.modality,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            fragments=fragments,
            raw_metadata={"sheets": wb.sheetnames, "sheet_count": sheet_count, "core": meta_dump},
            notes=notes,
        )
